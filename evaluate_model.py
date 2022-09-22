import argparse
import os

os.environ["CUDA_VISIBLE_DEVICES"] = "0"
import torch
from config import *
import numpy as np
config = Config()
from data.loader import data_loader
from models.Coll_Loss_one import CoLoss
from models.cvae_var import CVAE
from models.LSTM import LSTM, LSTMaddNoise
from utils import (
    relative_to_abs,
    get_dset_path,
    plot_multimodal,
    plot_best
)
from losses import displacement_error, final_displacement_error
from attrdict import AttrDict

# Change to your path here!
#DIR = home +'/Documents/col_aware/Your_project_name/checkpoint_with_model.pt'
DIR = os.path.join(config.home, config.checkpoint_start_from)


parser = argparse.ArgumentParser()
parser.add_argument('--model_path', default=DIR, type=str)
parser.add_argument('--num_samples', default=20, type=int)
parser.add_argument('--dset_type', default='test', type=str)
parser.add_argument('--dset', default="dataset", type=str)

seed = config.seed
torch.manual_seed(seed)
np.random.seed(seed)
# import locale
# locale.setlocale(locale.LC_ALL, 'de_DE.utf8')

def evaluate_helper(error, seq_start_end):
    sum_ = 0
    error = torch.stack(error, dim=1)
    ids = []
    for (start, end) in seq_start_end:
        start = start.item()
        end = end.item()
        _error = error[start:end]
        _error = torch.sum(_error, dim=0).unsqueeze(dim=1)
        _error, id = torch.min(_error, 0)
        ids.append(id.squeeze().item() )
        sum_ += _error.squeeze()
    return sum_, ids


def get_generator(checkpoint):
    model = CVAE()
    #in_size = 12
    #cond_size = 8
    #extra_layer_size = 400
    #latent_size = 2
    #model = CVAE(in_size, latent_size, cond_size, extra_layer_size)
    #model = LSTM()

    model.load_state_dict(checkpoint["best_state"])
    model.cuda()
    model.eval()
    return model


def cal_ade_fde(pred_traj_gt, pred_traj_fake, val_mask):
    ade = displacement_error(pred_traj_fake, pred_traj_gt, val_mask, mode="raw")
    fde = final_displacement_error(pred_traj_fake[-1], pred_traj_gt[-1], val_mask[-1], mode="raw")
    return ade, fde

def fast_coll_counter(pred_batch,seq_start_end, ids, mask):
    ids_of_col_szenes = np.zeros([seq_start_end.shape[0]])
    coll_pro_szene = 0
    stack_of_coll_indeces = []
    for i, (start, end) in enumerate(seq_start_end):
        start = start.item()
        end = end.item()
        if ids:
            pred = pred_batch[ids[i]]
        else:
            pred = pred_batch
        pred = pred
        currSzene = pred[:, start:end]
        dist_mat = torch.cdist(currSzene, currSzene, p=2.0, compute_mode='donot_use_mm_for_euclid_dist')
        dist_mat_one_triu = torch.triu(dist_mat)
        dist_mat_one_triu = dist_mat_one_triu * mask[:,start:end, start:end]
        filter_zeros = torch.logical_and(0. != dist_mat_one_triu,  dist_mat_one_triu <= config.collision_distance)
        filter_col_pos = torch.logical_and(0. != dist_mat,  dist_mat <= config.collision_distance)
        filter_zeros_sum=filter_zeros.sum().unsqueeze(dim=0).item()
        coll_pro_szene += filter_zeros_sum
        if filter_zeros_sum > 0.:
            ids_of_col_szenes[i] = 1
        stack_of_coll_indeces.append(filter_col_pos)
    count = len(seq_start_end) #* count_empty
    return coll_pro_szene, count, ids_of_col_szenes, stack_of_coll_indeces

def evaluate(args, loader, model, gt_coll=False, plot_traj=False):
    ade_outer, fde_outer = [], []
    total_traj = 0
    coll_pro_szenes_fake = 0.
    count_szenes_fake = 0.
    coll_pro_szenes = 0.
    count_szenes = 0.
    szene_id = 0.
    with torch.no_grad():
        for batch in loader:
            batch = [tensor.cuda() for tensor in batch]
            obs_traj, pred_traj_gt, obs_traj_rel, pred_traj_gt_rel, val_mask, \
             loss_mask, seq_start_end, nei_num_index, nei_num = batch
            ade, fde = [], []
            total_traj += nei_num.sum()
            batch_pred_traj_fake = []
            att_score_list_batch = []
            model_input = torch.cat((obs_traj_rel, pred_traj_gt_rel), dim=0)
            val_mask = val_mask[-config.pred_len :]
            for _ in range(args.num_samples):
                if plot_traj:
                    #COLOSS
                    #pred_traj_fake_rel = model(model_input, obs_traj, pred_traj_gt, seq_start_end, nei_num_index, nei_num)

                    # LSTM
                    #pred_traj_fake_rel = model(obs_traj_rel, pred_traj_gt_rel)

                    #CVAE
                    pred_traj_fake_rel, prior_mu, prior_logvar, recog_mu, recog_logvar, pred_mean, pred_var  = model(obs_traj_rel, pred_traj_gt_rel, obs_traj, nei_num_index, nei_num, testing=True)

                else:
                    #COLOSS
                    #pred_traj_fake_rel = model(model_input, obs_traj, pred_traj_gt, seq_start_end, nei_num_index, nei_num)

                    # LSTM
                    #pred_traj_fake_rel = model(obs_traj_rel, pred_traj_gt_rel)

                    #CVAE
                    pred_traj_fake_rel, prior_mu, prior_logvar, recog_mu, recog_logvar, pred_mean, pred_var  = model(obs_traj_rel, pred_traj_gt_rel, obs_traj, nei_num_index, nei_num, testing=True)

                pred_traj_fake_rel = pred_traj_fake_rel[-args.pred_len :]
                pred_traj_fake = relative_to_abs(pred_traj_fake_rel, obs_traj[-1])
                batch_pred_traj_fake.append(pred_traj_fake)
                ade_, fde_ = cal_ade_fde(pred_traj_gt, pred_traj_fake, val_mask)
                ade.append(ade_)
                fde.append(fde_)
            ade_sum, ids = evaluate_helper(ade, seq_start_end)
            fde_sum, _ = evaluate_helper(fde, seq_start_end)

            ade_outer.append(ade_sum)
            fde_outer.append(fde_sum)

            coll_pro_szene_fake, count_fake, ids_of_col_szenes_fake, stack_of_coll_indeces = fast_coll_counter(batch_pred_traj_fake,
                                                                                        seq_start_end, ids, nei_num_index)
            coll_pro_szenes_fake += coll_pro_szene_fake
            count_szenes_fake += count_fake

            # print(str(coll_pro_szene_fake) + " collisions in " + str(count_fake) + " scenes")
            # print("ids: ", ids)
            # if(coll_pro_szene_fake > 100):
            #     plot_best(obs_traj, pred_traj_gt, batch_pred_traj_fake, att_score_list_batch, ids,
            #                    stack_of_coll_indeces, seq_start_end,
            #                    ids_of_col_szenes_fake, loss_mask, config, szene_id)


            if(gt_coll):
                coll_pro_szene, count, ids_of_col_szenes, stack_of_coll_indeces = fast_coll_counter(
                    pred_traj_gt, seq_start_end, ids, nei_num_index)
                coll_pro_szenes += coll_pro_szene
                count_szenes += count

            if(plot_traj):
                plot_best(obs_traj, pred_traj_gt, batch_pred_traj_fake, att_score_list_batch, ids,
                               stack_of_coll_indeces, seq_start_end,
                               ids_of_col_szenes_fake, loss_mask, config, szene_id)
            szene_id += 1

        print(str(coll_pro_szenes_fake) + " colls total in " + str(count_szenes_fake) + " szenes")
        ade = sum(ade_outer) / (total_traj * args.pred_len)
        fde = sum(fde_outer) / (total_traj)
        act = coll_pro_szenes_fake / count_szenes_fake
        if gt_coll:
            act_gt = coll_pro_szenes / count_szenes
            return ade, fde, act, act_gt
        return ade, fde, act, 0

import openpyxl
def write_to_excel(args, ade, fde, act, act_gt):
    file = "evaluation_pc.xlsx"

    wb = openpyxl.load_workbook(filename=file)
    ws = wb['Sheet1']
    col = ws.max_column + 1
    new_col = [config.num_epochs, config.learning_rate, config.coeff_coll_loss, config.coeff_kldiv, config.coeff_nll,
               config.experiment_name, args.dset, ade.item(), fde.item(), act, act_gt]

    for row, entry in enumerate(new_col, start=1):
        print("write " + str(entry) + " to " + str(row) + ", " + str(col))
        ws.cell(row=row, column=col, value=entry)
    wb.save(file)

def main(args):
    if os.path.isdir(args.model_path):
        filenames = os.listdir(args.model_path)
        filenames.sort()
        paths = [
            os.path.join(args.model_path, file_) for file_ in filenames
        ]
    else:
        paths = [args.model_path]
        #print("paths: ", paths)
    for path in paths:
        print("path: ", path)
        checkpoint = torch.load(path)
        #print("checkpoint: ", checkpoint)
        generator = get_generator(checkpoint)
        _args = AttrDict(checkpoint['args'])
        path = get_dset_path(_args.dataset_name, args.dset_type)

        _, loader = data_loader(_args, path)
        ade, fde, act, act_gt = evaluate(_args, loader, generator, gt_coll=False, plot_traj=False)

        write_to_excel(args, ade, fde, act, act_gt)

        print(
            "Dataset:  {} \n"
            "ADE {:.4f} \n" 
            "FDE {:.4f} \n"
            "ACT {:.4f} \n".format(
                _args.dataset_name, ade, fde, act
            )
        )


if __name__ == "__main__":
    args = parser.parse_args()
    args.dset = config.dataset_name
    print("args: ", args)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False
    main(args)
